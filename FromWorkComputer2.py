import openpyxl
import os
import subprocess
import time

#wb = openpyxl.load_workbook('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\LSSJobsites_002_.xlsx')
#wbDummy=openpyxl.load_workbook('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\TestJobsites.xlsx')
wbDummyHome=openpyxl.load_workbook('C:\\Users\\ray\\Documents\\Projects\\Python\\TestJobsites.xlsx')


#sheet = wb.get_sheet_by_name('Sheet2')
#sheet=wbDummy.get_sheet_by_name('Sheet2')

sheet=wbDummyHome.get_sheet_by_name('Sheet2')

for i in range(2, 4, 1):
       # print(i, sheet.cell(row=i, column=1).value)
       # c=(i, sheet.cell(row=i, column=1).value)
       jobsiteid=(sheet.cell(row=i, column=1).value)
       print(jobsiteid)

# some code here writing to executable python file and close file
       file=open('executeFile.py','w')
       file.write('import requests\n')
       file.write('#url=http://wrenetdev:8081/MarthaAPI/api/Storage/'+ str(jobsiteid) + '/UploadDiagram?userID=1113\n')
       # add working request code
       file.write('#with open(fileString, "rb") as f:\n')
       file.write('#     r=requests.post(url,file={fileString, f})\n')
       file.write('print("Hello, World")')
       file.close

# execute python file
       # run above created file
       time.sleep(20)
       subprocess.call("executeFile.py", shell=True)
       # time delay and or confirm completion of file execution (try/catch block?)

# log that file has been processed


# This process will start again by virtue of top loop program using openpyx
