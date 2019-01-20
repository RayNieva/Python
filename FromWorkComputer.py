import requests
import openpyxl

#wb = openpyxl.load_workbook('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\LSSJobsites_002_.xlsx')
wbDummy=openpyxl.load_workbook('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\TestJobsites.xlsx')
#sheet = wb.get_sheet_by_name('Sheet2')
sheet=wbDummy.get_sheet_by_name('Sheet2')
for i in range(2, 4, 1):
      # print(i, sheet.cell(row=i, column=1).value)
       # c=(i, sheet.cell(row=i, column=1).value)
       jobsiteid=(sheet.cell(row=i, column=1).value)
       print(jobsiteid)
       url="http://wrenetdev:8081/MarthaAPI/api/Storage/"+ str(jobsiteid) + "/UploadDiagram?userID=1113"
       #fileString='C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\100.jpg'
       fileString='C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\' + str(jobsiteid) +'.jpg'
       print(fileString)
       with open(fileString, 'rb') as f:
           r=requests.post(url,file={fileString, f})
           print(r)

       with open('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\100.jpg', 'rb') as f:
           r=requests.post(url,file={'C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\100.jpg', f})
           print(r)




#url="http://wrenetdev:8081/MarthaAPI/api/Storage/100/UploadDiagram?userID=1113"
#with open('C:\\Users\\rnieva\\Documents\\Projects\\API2UploadDiagrams\\100.jpg', 'rb') as f:
    #r = requests.post(url, files = {'100.jpg' : f})

r

# This is weird changing back to http worked this time as before change to https from http made it worked.
